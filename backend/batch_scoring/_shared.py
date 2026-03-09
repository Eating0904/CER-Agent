import os
import re
import sys
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Django 環境初始化 ──────────────────────────────────────────────────────────


def setup_django():
    """初始化 Django 環境（需在 import Django model 之前呼叫）"""
    backend_path = Path(__file__).parent.parent
    sys.path.insert(0, str(backend_path))
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

    import django

    django.setup()


# ── 互動式選單 ─────────────────────────────────────────────────────────────────


def select_test_folder(base_dir: Path) -> Path:
    """
    列出 test_data 下的所有資料夾，讓使用者輸入編號選擇。

    Args:
        base_dir: mindmap/ 或 essay/ 的根目錄（含 test_data/ 子目錄）

    Returns:
        Path: 使用者選擇的資料夾路徑
    """
    test_data_dir = base_dir / 'data'

    if not test_data_dir.exists():
        print(f'❌ 找不到 test_data 目錄: {test_data_dir}')
        sys.exit(1)

    folders = sorted([d for d in test_data_dir.iterdir() if d.is_dir()])

    if not folders:
        print(f'❌ test_data 目錄下沒有任何資料夾: {test_data_dir}')
        sys.exit(1)

    EXCLUDED = {'article.txt', 'prompt.txt'}
    print('\n📂 可用的測試資料夾：')
    for idx, folder in enumerate(folders, 1):
        data_count = len([f for f in folder.iterdir() if f.is_file() and f.name not in EXCLUDED])
        has_prompt = '  ✦ 含自訂 prompt' if (folder / 'prompt.txt').exists() else ''
        print(f'  {idx}. {folder.name}  ({data_count} 筆資料){has_prompt}')

    while True:
        try:
            choice = input('\n請輸入資料夾編號: ').strip()
            idx = int(choice)
            if 1 <= idx <= len(folders):
                selected = folders[idx - 1]
                print(f'✅ 已選擇: {selected.name}')
                return selected
            else:
                print(f'⚠️  請輸入 1 到 {len(folders)} 之間的數字')
        except ValueError:
            print('⚠️  請輸入有效的數字')


# ── 資料讀取 ───────────────────────────────────────────────────────────────────


def load_article(folder_path: Path) -> str:
    """
    讀取資料夾下的 article.txt。

    Args:
        folder_path: 測試資料夾路徑

    Returns:
        str: 文章內容
    """
    article_path = folder_path / 'article.txt'
    if not article_path.exists():
        print(f'❌ 找不到 article.txt: {article_path}')
        sys.exit(1)

    with open(article_path, encoding='utf-8') as f:
        content = f.read().strip()

    if not content:
        print(f'⚠️  article.txt 是空的: {article_path}')

    return content


# ── Prompt 選擇 ────────────────────────────────────────────────────────────────


def select_prompt(folder_path: Path, agent) -> None:
    """
    偵測資料夾下是否有 prompt.txt。
    若有，詢問使用者要用內建還是自訂 prompt；
    若選自訂，讀取 prompt.txt 並覆蓋 agent.prompt_template。

    Args:
        folder_path: 測試資料夾路徑
        agent: Scoring Agent 實例（需有 prompt_template 屬性）
    """
    prompt_path = folder_path / 'prompt.txt'

    if not prompt_path.exists():
        print('📝 使用內建 Prompt')
        return

    print('\n📝 偵測到 prompt.txt，請選擇要使用的 Prompt：')
    print('  1. 使用內建 Prompt（backend 預設）')
    print('  2. 使用自訂 Prompt（prompt.txt）')

    while True:
        choice = input('\n請輸入選項 [1/2]: ').strip()
        if choice == '1':
            print('✅ 使用內建 Prompt')
            return
        elif choice == '2':
            with open(prompt_path, encoding='utf-8') as f:
                custom_prompt = f.read()
            agent.prompt_template = custom_prompt
            print('✅ 使用自訂 Prompt（prompt.txt）')
            return
        else:
            print('⚠️  請輸入 1 或 2')


# ── 評分輔助 ───────────────────────────────────────────────────────────────────


def extract_score_number(score_str) -> int | None:
    """
    從評分字串中提取數字分數。
    例如：「4分」→ 4，「3 分」→ 3，「4」→ 4

    Args:
        score_str: 評分字串或數字

    Returns:
        int | None: 提取到的整數分數，失敗則回傳 None
    """
    if isinstance(score_str, (int, float)):
        return int(score_str)

    cleaned = str(score_str).replace(' ', '').replace('分', '')
    match = re.search(r'\d+', cleaned)
    if match:
        return int(match.group())

    return None


# ── Excel 共用樣式 ─────────────────────────────────────────────────────────────


def get_excel_styles() -> dict:
    """
    回傳 Excel 的共用樣式 dict，供各腳本使用。

    Returns:
        dict: 包含 base_font, header_font, center_align, left_align,
              thin_border, header_fill_gray 的樣式 dict
    """
    return {
        'base_font': Font(name='微軟正黑體', size=11),
        'header_font': Font(name='微軟正黑體', bold=True, size=11),
        'center_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left_align': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        ),
        'header_fill_gray': PatternFill(
            start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'
        ),
    }


# ── Excel Cell 寫入 ────────────────────────────────────────────────────────────


def write_cell(cell, value, font, alignment, border, fill=None) -> None:
    """
    單一 cell 的屬性一次設定完畢。

    Args:
        cell: openpyxl Cell 物件
        value: 儲存格內容
        font: Font 物件
        alignment: Alignment 物件
        border: Border 物件
        fill: PatternFill 物件（選填，不傳則不設底色）
    """
    cell.value = value
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill is not None:
        cell.fill = fill
