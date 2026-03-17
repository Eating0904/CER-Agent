"""
batch_scoring/batch_scoring_mindmap.py

Mindmap CER 評分批次測試腳本。

直接呼叫 ScoringAgent 進行評分（繞過前端 chat 流程），
並將所有評分結果匯出為 Excel。

執行方式（從 backend/ 目錄）：
    uv run python batch_scoring/batch_scoring_mindmap.py
"""

import json
import sys
from datetime import datetime
from pathlib import Path

from langchain_core.messages import HumanMessage
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 載入共用模組
sys.path.insert(0, str(Path(__file__).parent))
from _shared import (
    extract_score_number,
    get_excel_styles,
    load_article,
    select_prompt,
    select_test_folder,
    setup_django,
    write_cell,
)

# 初始化 Django 環境
setup_django()

from apps.chatbot.langgraph.mindmap.agents.scoring_agent import ScoringAgent  # noqa: E402

# ── 資料讀取 ──────────────────────────────────────────────────────────────────


def load_json_files(folder_path: Path) -> list[tuple[str, dict]]:
    """
    讀取資料夾下所有的 .json 檔（依檔名排序）。

    Args:
        folder_path: 測試資料夾路徑

    Returns:
        list[tuple[str, dict]]: [(檔名 stem, 資料), ...] 的清單
    """
    json_files = sorted(folder_path.glob('*.json'))

    if not json_files:
        print(f'❌ 資料夾下沒有任何 JSON 檔: {folder_path}')
        sys.exit(1)

    results = []
    for json_path in json_files:
        with open(json_path, encoding='utf-8') as f:
            data = json.load(f)
        results.append((json_path.stem, data))

    return results


# ── 顏色定義 ──────────────────────────────────────────────────────────────────

# 依分數的底色
SCORE_FILLS = {
    5: PatternFill(start_color='FFD9EC', end_color='FFD9EC', fill_type='solid'),  # 粉
    4: PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),  # 藍
    3: PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),  # 橘
    2: PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # 綠
    1: PatternFill(start_color='E8D5F5', end_color='E8D5F5', fill_type='solid'),  # 紫
    0: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # 黃
}

DIMENSIONS = [
    'C_覆蓋',
    'C_細節',
    'E_內容',
    'E_連線',
    'R_合理',
]
METADATA_KEYS = {
    'C_覆蓋': ('claim_coverage_score', 'claim_coverage_feedback'),
    'C_細節': ('claim_precision_score', 'claim_precision_feedback'),
    'E_內容': ('evidence_coverage_score', 'evidence_coverage_feedback'),
    'E_連線': ('evidence_connection_score', 'evidence_connection_feedback'),
    'R_合理': ('reasoning_score', 'reasoning_feedback'),
}


# ── 評分邏輯 ──────────────────────────────────────────────────────────────────


def score_one(file_id: str, data: dict, article_content: str, agent: ScoringAgent) -> dict:
    """
    對單筆 mindmap 資料進行評分。

    Args:
        file_id: JSON 檔名（不含副檔名）
        data: JSON 資料（含 nodes / edges）
        article_content: 文章內容
        agent: ScoringAgent 實例

    Returns:
        dict: 包含 id、各維度 score 和 feedback 的結果
    """
    mind_map_data = {
        'nodes': data.get('nodes', []),
        'edges': data.get('edges', []),
    }

    messages = agent.prepare_messages(
        messages=[HumanMessage(content=json.dumps({'context': {'mind_map_data': mind_map_data}}))],
        article_content=article_content,
    )

    response = agent.llm.invoke(messages)
    metadata = agent.extract_metadata(response)

    result = {'id': file_id}
    for dim in DIMENSIONS:
        score_key, feedback_key = METADATA_KEYS[dim]
        result[f'{dim}_score'] = extract_score_number(metadata.get(score_key, ''))
        result[f'{dim}_feedback'] = metadata.get(feedback_key, '')

    return result


# ── Excel 輸出 ────────────────────────────────────────────────────────────────


def generate_excel(results: list[dict], output_path: Path) -> None:
    """
    產生 Mindmap 評分結果 Excel。

    欄位：A=ID | B=C_覆蓋分數 | C=C_細節分數 | D=E_內容分數 | E=E_連線分數 | F=R_合理分數 |
        G=C_覆蓋 Feedback | H=C_細節 Feedback | I=E_內容 Feedback | J=E_連線 Feedback | K=R_合理 Feedback

    Args:
        results: score_one() 回傳的結果清單
        output_path: 輸出檔案路徑
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mindmap CER 評分結果'

    styles = get_excel_styles()
    base_font = styles['base_font']
    header_font = styles['header_font']
    center_align = styles['center_align']
    left_align = styles['left_align']
    thin_border = styles['thin_border']

    score_cols = {
        'C_覆蓋': 2,
        'C_細節': 3,
        'E_內容': 4,
        'E_連線': 5,
        'R_合理': 6,
    }
    feedback_cols = {
        'C_覆蓋': 7,
        'C_細節': 8,
        'E_內容': 9,
        'E_連線': 10,
        'R_合理': 11,
    }

    # ── 表頭（第 1 列）────────────────────────────────────────────────────────
    write_cell(ws.cell(row=1, column=1), 'ID', header_font, center_align, thin_border)

    for dim in DIMENSIONS:
        write_cell(
            ws.cell(row=1, column=score_cols[dim]),
            f'{dim} 分數',
            header_font,
            center_align,
            thin_border,
        )
        write_cell(
            ws.cell(row=1, column=feedback_cols[dim]),
            f'{dim} Feedback',
            header_font,
            center_align,
            thin_border,
        )

    # ── 資料列（從第 2 列開始）────────────────────────────────────────────────
    for row_num, result in enumerate(results, 2):
        write_cell(
            ws.cell(row=row_num, column=1), result['id'], base_font, center_align, thin_border
        )

        for dim in DIMENSIONS:
            score_val = result[f'{dim}_score']
            fill = SCORE_FILLS.get(score_val)
            write_cell(
                ws.cell(row=row_num, column=score_cols[dim]),
                score_val,
                base_font,
                center_align,
                thin_border,
                fill,
            )
            write_cell(
                ws.cell(row=row_num, column=feedback_cols[dim]),
                result[f'{dim}_feedback'],
                base_font,
                left_align,
                thin_border,
                fill,
            )

    # ── 欄寬 & 凍結窗格 ───────────────────────────────────────────────────────
    ws.freeze_panes = 'G2'  # 凍結第 1 列 + A~F 欄（ID 和五個分數）
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 55
    ws.column_dimensions['H'].width = 55
    ws.column_dimensions['I'].width = 55
    ws.column_dimensions['J'].width = 55
    ws.column_dimensions['K'].width = 55

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'✅ Excel 已儲存至: {output_path}')


# ── 主程式 ────────────────────────────────────────────────────────────────────


def main():
    base_dir = Path(__file__).parent / 'mindmap'

    print('=' * 60)
    print('  Mindmap CER 評分批次測試')
    print('=' * 60)

    # 選擇測試資料夾
    folder = select_test_folder(base_dir)

    # 讀取資料
    print('\n📂 載入測試資料...')
    article_content = load_article(folder)
    json_files = load_json_files(folder)
    print(f'✅ 共 {len(json_files)} 筆測試資料')

    # 初始化 Agent
    print('\n🤖 初始化 ScoringAgent...')
    agent = ScoringAgent()
    print('✅ 初始化完成')

    # 選擇 Prompt
    select_prompt(folder, agent)

    # 執行評分
    print(f'\n🧪 開始評分（共 {len(json_files)} 筆）...')
    results = []

    for idx, (file_id, data) in enumerate(json_files, 1):
        print(f'  [{idx}/{len(json_files)}] 評分中: {file_id}')
        try:
            result = score_one(file_id, data, article_content, agent)
            results.append(result)
            scores = ' '.join([f'{dim}={result[f"{dim}_score"]}' for dim in DIMENSIONS])
            print(f'         → {scores}')
        except Exception as e:
            print(f'         ❌ 失敗: {str(e)}')
            error_result: dict = {'id': file_id}
            for dim in DIMENSIONS:
                error_result[f'{dim}_score'] = None
                error_result[f'{dim}_feedback'] = f'錯誤: {str(e)}' if dim == DIMENSIONS[0] else ''
            results.append(error_result)

    # 輸出 Excel
    print('\n📊 產生 Excel 報告...')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = base_dir / 'results' / folder.name / f'{timestamp}.xlsx'
    generate_excel(results, output_path)

    print('\n' + '=' * 60)
    print(f'完成！共評分 {len(results)} 筆')
    print('=' * 60)


if __name__ == '__main__':
    main()
