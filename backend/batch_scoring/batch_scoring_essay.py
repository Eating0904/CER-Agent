"""
batch_scoring/batch_scoring_essay.py

Essay 評分批次測試腳本。

直接呼叫 EssayScoringAgent 進行評分（繞過前端 chat 流程），
並將所有評分結果匯出為 Excel。

執行方式（從 backend/ 目錄）：
    uv run python batch_scoring/batch_scoring_essay.py
"""

import json
import sys
from datetime import datetime
from pathlib import Path

from langchain_core.messages import HumanMessage
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 載入共用模組
sys.path.insert(0, str(Path(__file__).parent))
from _shared import (
    extract_score_number,
    get_excel_styles,
    load_article,
    select_prompt,
    select_test_folder,
    setup_django,
    write_cell,
)

# 初始化 Django 環境
setup_django()

from apps.chatbot.langgraph.essay.agents.scoring_agent import EssayScoringAgent  # noqa: E402

# ── 維度定義 ──────────────────────────────────────────────────────────────────

DIMENSIONS = [
    'explanation_of_issues',
    'evidence_integration',
    'influence_of_context',
    'students_position',
    'conclusions',
]

# 對應 extract_metadata 的 key 名稱
METADATA_KEYS = {dim: (f'{dim}_score', f'{dim}_feedback') for dim in DIMENSIONS}

# 表頭顯示名稱
DIMENSION_LABELS = {
    'explanation_of_issues': 'Explanation of Issues',
    'evidence_integration': 'Evidence Integration',
    'influence_of_context': 'Influence of Context',
    'students_position': "Student's Position",
    'conclusions': 'Conclusions',
}

# 依分數的底色
SCORE_FILLS = {
    4: PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),  # 藍
    3: PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),  # 橘
    2: PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # 綠
    1: PatternFill(start_color='E8D5F5', end_color='E8D5F5', fill_type='solid'),  # 紫
    0: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # 黃
}


# ── 資料讀取 ────────────────────────────────────────────────────────────────


def load_essay_files(folder_path: Path) -> list[tuple[str, str]]:
    """
    讀取資料夾下所有 .txt 檔（排除 article.txt 和 prompt.txt）。

    Args:
        folder_path: 測試資料夾路徑

    Returns:
        list[tuple[str, str]]: [(檔名 stem, essay 內容), ...] 的清單（依檔名排序）
    """
    excluded = {'article', 'prompt'}
    txt_files = sorted(f for f in folder_path.glob('*.txt') if f.stem not in excluded)

    if not txt_files:
        print(f'❌ 資料夾下沒有任何 .txt 檔: {folder_path}')
        sys.exit(1)

    results = []
    for txt_path in txt_files:
        with open(txt_path, encoding='utf-8') as f:
            content = f.read().strip()
        results.append((txt_path.stem, content))

    return results


# ── 評分邏輯 ──────────────────────────────────────────────────────────────────


def score_one(
    file_id: str, essay_content: str, article_content: str, agent: EssayScoringAgent
) -> dict:
    """
    對單筆 essay 內容進行評分。

    Args:
        file_id: 檔名（不含副檔名）
        essay_content: essay 文字內容
        article_content: 閱讀文章內容
        agent: EssayScoringAgent 實例

    Returns:
        dict: 包含 id、各維度 score 和 feedback 的結果
    """
    messages = agent.prepare_messages(
        messages=[HumanMessage(content=json.dumps({'context': {'essay_content': essay_content}}))],
        article_content=article_content,
    )

    response = agent.llm.invoke(messages)
    metadata = agent.extract_metadata(response)

    result = {'id': file_id}
    for dim in DIMENSIONS:
        score_key, feedback_key = METADATA_KEYS[dim]
        result[f'{dim}_score'] = extract_score_number(metadata.get(score_key, ''))
        result[f'{dim}_feedback'] = metadata.get(feedback_key, '')

    return result


# ── Excel 輸出 ────────────────────────────────────────────────────────────────


def generate_excel(results: list[dict], output_path: Path) -> None:
    """
    產生 Essay 評分結果 Excel。

    欄位：
      A = ID
      B~F = 5 個維度的分數
      G~K = 5 個維度的 feedback

    Args:
        results: score_one() 回傳的結果清單
        output_path: 輸出檔案路徑
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Essay 評分結果'

    styles = get_excel_styles()
    base_font = styles['base_font']
    header_font = styles['header_font']
    center_align = styles['center_align']
    left_align = styles['left_align']
    thin_border = styles['thin_border']

    # 欄位對應
    score_col = {dim: 2 + i for i, dim in enumerate(DIMENSIONS)}  # B~F
    feedback_col = {dim: 7 + i for i, dim in enumerate(DIMENSIONS)}  # G~K

    # ── 表頭（第 1 列）────────────────────────────────────────────────────────
    write_cell(ws.cell(row=1, column=1), 'ID', header_font, center_align, thin_border)

    for dim in DIMENSIONS:
        write_cell(
            ws.cell(row=1, column=score_col[dim]),
            f'{DIMENSION_LABELS[dim]}\n分數',
            header_font,
            center_align,
            thin_border,
        )
        write_cell(
            ws.cell(row=1, column=feedback_col[dim]),
            f'{DIMENSION_LABELS[dim]} Feedback',
            header_font,
            center_align,
            thin_border,
        )

    # ── 資料列（從第 2 列開始）────────────────────────────────────────────────
    for row_num, result in enumerate(results, 2):
        write_cell(
            ws.cell(row=row_num, column=1), result['id'], base_font, center_align, thin_border
        )

        for dim in DIMENSIONS:
            score_val = result[f'{dim}_score']
            fill = SCORE_FILLS.get(score_val)
            write_cell(
                ws.cell(row=row_num, column=score_col[dim]),
                score_val,
                base_font,
                center_align,
                thin_border,
                fill,
            )
            write_cell(
                ws.cell(row=row_num, column=feedback_col[dim]),
                result[f'{dim}_feedback'],
                base_font,
                left_align,
                thin_border,
                fill,
            )

    # ── 欄寬 & 凍結窗格 ───────────────────────────────────────────────────────
    ws.freeze_panes = 'G2'  # 凍結第 1 列 + A~F 欄
    ws.row_dimensions[1].height = 50
    ws.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 14
    for col_letter in ['G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 55

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'✅ Excel 已儲存至: {output_path}')


# ── 主程式 ────────────────────────────────────────────────────────────────────


def main():
    base_dir = Path(__file__).parent / 'essay'

    print('=' * 60)
    print('  Essay 評分批次測試')
    print('=' * 60)

    # 選擇測試資料夾
    folder = select_test_folder(base_dir)

    # 讀取資料
    print('\n📂 載入測試資料...')
    article_content = load_article(folder)
    essay_files = load_essay_files(folder)
    print(f'✅ 共 {len(essay_files)} 筆測試資料')

    # 初始化 Agent
    print('\n🤖 初始化 EssayScoringAgent...')
    agent = EssayScoringAgent()
    print('✅ 初始化完成')

    # 選擇 Prompt
    select_prompt(folder, agent)

    # 執行評分
    print(f'\n🧪 開始評分（共 {len(essay_files)} 筆）...')
    results = []

    for idx, (file_id, essay_content) in enumerate(essay_files, 1):
        print(f'  [{idx}/{len(essay_files)}] 評分中: {file_id}')
        try:
            result = score_one(file_id, essay_content, article_content, agent)
            results.append(result)
            scores = ' | '.join(
                f'{DIMENSION_LABELS[d][:4]}={result[f"{d}_score"]}' for d in DIMENSIONS
            )
            print(f'         → {scores}')
        except Exception as e:
            print(f'         ❌ 失敗: {str(e)}')
            results.append(
                {
                    'id': file_id,
                    **{f'{dim}_score': None for dim in DIMENSIONS},
                    **{
                        f'{dim}_feedback': (f'錯誤: {str(e)}' if dim == DIMENSIONS[0] else '')
                        for dim in DIMENSIONS
                    },
                }
            )

    # 輸出 Excel
    print('\n📊 產生 Excel 報告...')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = base_dir / 'results' / folder.name / f'{timestamp}.xlsx'
    generate_excel(results, output_path)

    print('\n' + '=' * 60)
    print(f'完成！共評分 {len(results)} 筆')
    print('=' * 60)


if __name__ == '__main__':
    main()
