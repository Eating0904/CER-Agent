import json
import os
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.config import DB_CONFIG, get_user_class
from src.models import UserActionRecord
from src.time_utils import to_taiwan_str, ts_to_filename_str, iter_slots, get_now_tw_str
from src.langfuse_trace import extract_trace_data, TARGET_ACTION_TYPES

FONT_NAME = "Calibri"
FONT_SIZE = 12
COL_WIDTHS = [9, 25, 82, 9, 9, 15, 15, 10, 10, 21]
HEADER_COLOR = "BFBFBF"

_COLOR_GROUPS: dict[str, list[str]] = {
    # 紫
    "AAAAFF": ["ai_feedback_shown"],
    # 藍
    "D1E4FF": [
        "add_node",
        "delete_node",
        "node_edit_start",
        "node_edit_end",
        "add_edge",
        "delete_edge",
    ],
    # 粉
    "FFC1E0": ["click_feedback_ask", "chat_in_mindmap", "chat_in_essay"],
    # 黃
    "FFF2CC": [
        "click_add_new_button",
        "switch_map",
        "login",
        "create_map",
        "rename_map",
    ],
    # 綠
    "D9EAD3": ["essay_edit_start", "essay_edit_end"],
    # 橘
    "FF8F59": ["click_scoring_mindmap", "click_scoring_essay"],
    # 紅
    "CE0000": ["paste_detected"],
}
ACTION_TYPE_COLORS: dict[str, str] = {
    at: color for color, actions in _COLOR_GROUPS.items() for at in actions
}

PAGE_VIEW_END_COLORS: dict[str, str] = {
    "article": "FFF2CC",
    "mindmap": "D1E4FF",
    "essay": "D9EAD3",
}


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _make_font(bold: bool = False) -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold)


def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def fetch_user_ids(conn) -> list[int]:
    with conn.cursor() as cur:
        cur.execute("SELECT DISTINCT user_id FROM user_action ORDER BY user_id;")
        return [row[0] for row in cur.fetchall()]


def fetch_actions(
    conn,
    user_id: int,
    start_dt: datetime | None = None,
    end_dt: datetime | None = None,
) -> list[UserActionRecord]:
    query = """
        SELECT id, user_id, action_type, timestamp, map_id, essay_id, metadata
        FROM user_action
        WHERE user_id = %s
    """
    params: list = [user_id]

    if start_dt is not None:
        query += " AND timestamp >= %s"
        params.append(start_dt)
    if end_dt is not None:
        query += " AND timestamp <= %s"
        params.append(end_dt)

    query += " ORDER BY timestamp;"

    with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute(query, params)
        return [UserActionRecord(**dict(row)) for row in cur.fetchall()]


def export_user_excel(
    user_id: int,
    slot_data: list[tuple[str, list[UserActionRecord]]],
    start_label: str,
    end_label: str,
    output_base: str = "user_action",
) -> str:
    user_class = get_user_class(user_id)
    formatted_id = str(user_id).zfill(2) if user_id < 10 else str(user_id)
    dir_path = os.path.join(output_base, user_class, str(formatted_id))
    os.makedirs(dir_path, exist_ok=True)

    filepath = os.path.join(
        dir_path, f"{formatted_id}-{start_label}_to_{end_label}.xlsx"
    )

    wb = Workbook()

    for idx, (slot_label, rows) in enumerate(slot_data):
        if idx == 0:
            ws = wb.active
            ws.title = slot_label
        else:
            ws = wb.create_sheet(title=slot_label)

        headers = [
            "id",
            "action_type",
            "metadata",
            "map_id",
            "essay_id",
            "query",
            "final_response",
            "input",
            "output",
            "timestamp",
        ]
        ws.append(headers)

        header_fill = _make_fill(HEADER_COLOR)
        header_font = _make_font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        data_font = _make_font()
        center_align = Alignment(horizontal="center")

        def _to_cell(value) -> str:
            if value is None:
                return ""
            if isinstance(value, dict):
                return json.dumps(value, ensure_ascii=False)
            return str(value)

        for row in rows:
            action_type = row.action_type or ""

            trace_data = None
            if action_type in TARGET_ACTION_TYPES:
                trace_id = None
                if isinstance(row.metadata, dict):
                    trace_id = row.metadata.get("langfuse_trace_id")
                if trace_id:
                    trace_data = extract_trace_data(trace_id)

            row_values = [
                row.id,
                action_type,
                (
                    json.dumps(row.metadata, ensure_ascii=False)
                    if row.metadata is not None
                    else ""
                ),
                row.map_id,
                row.essay_id,
            ]
            if trace_data:
                row_values.extend(
                    [
                        trace_data.get("query", ""),
                        trace_data.get("final_response", ""),
                        _to_cell(trace_data.get("llm_input")),
                        _to_cell(trace_data.get("llm_output")),
                    ]
                )
            else:
                row_values.extend(["", "", "", ""])
            row_values.append(to_taiwan_str(row.timestamp))

            ws.append(row_values)

            cur_row = ws.max_row

            color = ACTION_TYPE_COLORS.get(action_type)
            if color:
                ws.cell(row=cur_row, column=2).fill = _make_fill(color)

            if action_type == "page_view_end":
                view_val = (
                    row.metadata.get("view", "")
                    if isinstance(row.metadata, dict)
                    else ""
                )
                b_color = PAGE_VIEW_END_COLORS.get(view_val)
                if b_color:
                    ws.cell(row=cur_row, column=3).fill = _make_fill(b_color)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=cur_row, column=col_idx)
                cell.font = data_font
                if col_idx in (1, 4, 5):
                    cell.alignment = center_align

        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "B2"

    wb.save(filepath)
    return filepath


def export_all(
    user_ids: list[int],
    start_dt: datetime,
    end_dt: datetime,
    output_base: str = "user_action",
) -> None:
    conn = get_connection()

    label_start = ts_to_filename_str(start_dt)
    label_end = ts_to_filename_str(end_dt)

    os.makedirs(output_base, exist_ok=True)
    now_str = get_now_tw_str()
    log_path = os.path.join(output_base, f"{now_str}.txt")

    with open(log_path, "w", encoding="utf-8") as log_f:
        try:
            total = len(user_ids)
            slots = list(iter_slots(start_dt, end_dt))

            for i, uid in enumerate(user_ids, start=1):
                log_f.write(f"\n  [{i}/{total}] 處理 user_id={uid} ...\n")

                user_slot_data: list[tuple[str, list[UserActionRecord]]] = []
                for slot_label, slot_start, slot_end in slots:
                    rows = fetch_actions(conn, uid, slot_start, slot_end)

                    if not rows:
                        slot_label = f"X_{slot_label}"

                    user_slot_data.append((slot_label, rows))
                    status = f"({len(rows)} 筆)" if rows else "(無資料，僅標題列)"
                    log_f.write(f"    - {slot_label} {status}\n")

                export_user_excel(
                    uid, user_slot_data, label_start, label_end, output_base
                )
        finally:
            conn.close()
